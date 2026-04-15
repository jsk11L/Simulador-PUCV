from pathlib import Path
import json
import re

from openpyxl import load_workbook
from pypdf import PdfReader

base = Path(r"c:\Users\psofi\OneDrive\Desktop\cosas\proyects\Simulador PUCV\Simulador-PUCV\original")
xlsx_path = base / "Civilelectrica93.xlsx"
pdf_path = base / "simulacion-del-desempeno-de-los-estudiantes-en-el-plan-de-estudio-de-la-carrera-de-ingenieria-civil-electrica (1).pdf"

wb = load_workbook(xlsx_path, data_only=True)
print("=== XLSX SHEETS ===")
print(json.dumps(wb.sheetnames, ensure_ascii=False))

ws = wb["Malla"]
print("=== MALLA HEADER ===")
print([ws.cell(1, c).value for c in range(1, 12)])
print("=== MALLA SAMPLE ===")
for r in range(2, 12):
    print([ws.cell(r, c).value for c in range(1, 12)])

print("=== PROGRAMACION COUNTS ===")
for sheet in ["ProgramacionB", "ProgramacionPE", "ProgramacionP", "ProgramacionS"]:
    wp = wb[sheet]
    left = [wp.cell(r, 1).value for r in range(2, wp.max_row + 1) if wp.cell(r, 1).value is not None]
    right = [wp.cell(r, 2).value for r in range(2, wp.max_row + 1) if wp.cell(r, 2).value is not None]
    print(sheet, len(left), len(right), "sample", left[:5], right[:5])

base_malla = wb["Malla"]
prog_b = wb["ProgramacionB"]
first_sem = {
    int(prog_b.cell(r, 1).value)
    for r in range(2, prog_b.max_row + 1)
    if isinstance(prog_b.cell(r, 1).value, (int, float))
}
second_sem = {
    int(prog_b.cell(r, 2).value)
    for r in range(2, prog_b.max_row + 1)
    if isinstance(prog_b.cell(r, 2).value, (int, float))
}

simulapucv_rows = []
for r in range(2, base_malla.max_row + 1):
    semestre = base_malla.cell(r, 1).value
    sigla = base_malla.cell(r, 2).value
    if sigla is None:
        continue

    sigla_num = int(sigla)
    creditos = int(base_malla.cell(r, 3).value or 0)
    reprob = float(base_malla.cell(r, 4).value or 0)
    nreq = int(base_malla.cell(r, 5).value or 0)

    reqs = []
    for c in range(6, 12):
        v = base_malla.cell(r, c).value
        if isinstance(v, (int, float)) and int(v) != 0:
            reqs.append(str(int(v)))
    if nreq > 0:
        reqs = reqs[:nreq]

    dictacion = "semestral" if (sigla_num in first_sem and sigla_num in second_sem) else "anual"

    simulapucv_rows.append(
        {
            "id": str(sigla_num),
            "cred": creditos,
            "rep": round(reprob, 2),
            "reqs": reqs,
            "semestre": int(semestre),
            "dictacion": dictacion,
        }
    )

out_json = base / "_malla_simulapucv_base.json"
out_json.write_text(json.dumps(simulapucv_rows, ensure_ascii=False, indent=2), encoding="utf-8")
print("=== MALLA EXPORT ===")
print("COUNT", len(simulapucv_rows), "FILE", out_json)

pdf_text = "\n".join((page.extract_text() or "") for page in PdfReader(str(pdf_path)).pages)
print("=== PDF METRIC KEYWORD COUNTS ===")
for k in ["PPE", "PSCE", "EE", "PEO", "R-10", "R-10Mat", "R-10>40", "4AS", "PF", "CAS", "PE", "actual", "tabla"]:
    print(k, pdf_text.lower().count(k.lower()))

print("=== PDF LINES WITH SCENARIOS OR METRICS ===")
for line in pdf_text.splitlines():
    s = line.strip()
    if not s:
        continue
    if re.search(r"(Actual|CAS|R-10|R\+10|R-10Mat|R-10>40|4AS|PF|CI|PPE|PSCE|PEO|EE)", s, flags=re.IGNORECASE):
        if len(s) <= 180:
            print(s)
