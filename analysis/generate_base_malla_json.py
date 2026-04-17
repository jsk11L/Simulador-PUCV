from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    xlsx_path = root / "original" / "Civilelectrica93.xlsx"
    out_path = root / "original" / "_malla_simulapucv_base.json"

    wb = load_workbook(xlsx_path, data_only=True)
    malla = wb["Malla"]
    prog_b = wb["ProgramacionB"]

    first_sem = {
        int(prog_b.cell(r, 1).value)
        for r in range(2, prog_b.max_row + 1)
        if isinstance(prog_b.cell(r, 1).value, (int, float))
    }
    second_sem = {
        int(prog_b.cell(r, 2).value)
        for r in range(2, prog_b.max_row + 1)
        if isinstance(prog_b.cell(r, 2).value, (int, float))
    }

    rows = []
    for r in range(2, malla.max_row + 1):
        semestre = malla.cell(r, 1).value
        sigla = malla.cell(r, 2).value
        if sigla is None:
            continue

        sigla_num = int(sigla)
        cred = int(malla.cell(r, 3).value or 0)
        rep = float(malla.cell(r, 4).value or 0)
        nreq = int(malla.cell(r, 5).value or 0)

        reqs = []
        for c in range(6, 12):
            v = malla.cell(r, c).value
            if isinstance(v, (int, float)) and int(v) != 0:
                reqs.append(str(int(v)))
        if nreq > 0:
            reqs = reqs[:nreq]

        dictacion = "semestral" if (sigla_num in first_sem and sigla_num in second_sem) else "anual"

        rows.append(
            {
                "id": str(sigla_num),
                "cred": cred,
                "rep": round(rep, 2),
                "reqs": reqs,
                "semestre": int(semestre),
                "dictacion": dictacion,
            }
        )

    out_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} rows to {out_path}")


if __name__ == "__main__":
    main()
