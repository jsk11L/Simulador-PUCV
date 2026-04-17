from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


SCENARIOS = [
    ("caso_actual", "Caso Actual", "Malla", "ProgramacionB"),
    ("pe", "PE", "Malla", "ProgramacionPE"),
    ("cas", "CAS", "Malla", "ProgramacionS"),
    ("r_10", "R-10", "Malla10me", "ProgramacionB"),
    ("r_mas_10", "R+10", "Malla10ma", "ProgramacionB"),
    ("r_10_gt_40", "R-10>40", "MallaR1050", "ProgramacionB"),
    ("pf", "PF", "MallaPF", "ProgramacionB"),
]


def read_malla(ws):
    first_sem = set()
    second_sem = set()

    rows = []
    for r in range(2, ws.max_row + 1):
        semestre = ws.cell(r, 1).value
        sigla = ws.cell(r, 2).value
        if sigla is None:
            continue

        sigla_num = int(sigla)
        cred = int(ws.cell(r, 3).value or 0)
        rep = float(ws.cell(r, 4).value or 0)
        nreq = int(ws.cell(r, 5).value or 0)

        reqs = []
        for c in range(6, 12):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and int(v) != 0:
                reqs.append(str(int(v)))
        if nreq > 0:
            reqs = reqs[:nreq]

        if isinstance(semestre, (int, float)):
            semestre = int(semestre)
        else:
            semestre = 0

        rows.append(
            {
                "id": str(sigla_num),
                "cred": cred,
                "rep": round(rep, 2),
                "reqs": reqs,
                "semestre": semestre,
                "dictacion": "",
            }
        )

    return rows


def read_programacion(ws):
    impar = []
    par = []

    for r in range(2, ws.max_row + 1):
        left = ws.cell(r, 1).value
        right = ws.cell(r, 2).value
        if isinstance(left, (int, float)):
            impar.append(str(int(left)))
        if isinstance(right, (int, float)):
            par.append(str(int(right)))

    return {"impar": impar, "par": par}


def infer_dictacion(rows, programacion):
    odd = set(programacion["impar"])
    even = set(programacion["par"])
    for row in rows:
        sigla = row["id"]
        in_odd = sigla in odd
        in_even = sigla in even
        row["dictacion"] = "semestral" if (in_odd and in_even) else "anual"


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    xlsx_path = root / "original" / "Civilelectrica93.xlsx"
    out_dir = root / "original" / "scenarios"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)

    for scenario_id, label, malla_sheet, prog_sheet in SCENARIOS:
        malla_ws = wb[malla_sheet]
        prog_ws = wb[prog_sheet]

        asignaturas = read_malla(malla_ws)
        programacion = read_programacion(prog_ws)
        infer_dictacion(asignaturas, programacion)

        out_path = out_dir / f"{scenario_id}.json"
        out_path.write_text(
            json.dumps(
                {
                    "scenario": label,
                    "asignaturas": asignaturas,
                    "programacion": programacion,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
