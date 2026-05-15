from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def infer_group(column_name: str) -> str:
    name = column_name.lower()
    if name == "scenario":
        return "scenario"
    if name.startswith("paper_"):
        return "paper"
    if name.startswith("go_"):
        return "go"
    if name.startswith("matlab_"):
        return "matlab"
    if name.startswith("delta_"):
        return "delta"
    return "other"


def reorder_columns(headers: list[str]) -> list[str]:
    scenario = [h for h in headers if h == "scenario"]

    metric_triplets = [
        ("paper_ppe", "go_ppe", "matlab_ppe"),
        ("paper_psce", "go_psce", "matlab_psce"),
        ("paper_ee", "go_ee", "matlab_ee"),
        ("paper_peo", "go_peo", "matlab_peo"),
    ]

    ordered = []
    ordered.extend(scenario)
    for triplet in metric_triplets:
        ordered.extend([h for h in triplet if h in headers])

    deltas = [h for h in headers if h.startswith("delta_")]
    ordered.extend(deltas)

    # Add any remaining columns (defensive)
    used = set(ordered)
    ordered.extend([h for h in headers if h not in used])

    return ordered


def apply_styles(ws, headers: list[str]) -> None:
    fill_paper = PatternFill(fill_type="solid", fgColor="C8E6C9")
    fill_go = PatternFill(fill_type="solid", fgColor="BBDEFB")
    fill_matlab = PatternFill(fill_type="solid", fgColor="FFCDD2")
    fill_delta = PatternFill(fill_type="solid", fgColor="EEEEEE")
    fill_scenario = PatternFill(fill_type="solid", fgColor="E8EAF6")

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = center

        group = infer_group(header)
        if group == "paper":
            cell.fill = fill_paper
        elif group == "go":
            cell.fill = fill_go
        elif group == "matlab":
            cell.fill = fill_matlab
        elif group == "delta":
            cell.fill = fill_delta
        elif group == "scenario":
            cell.fill = fill_scenario

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = headers[cell.column - 1]
            group = infer_group(header)

            if group == "paper":
                cell.fill = fill_paper
            elif group == "go":
                cell.fill = fill_go
            elif group == "matlab":
                cell.fill = fill_matlab
            elif group == "delta":
                cell.fill = fill_delta

            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
                cell.alignment = center

    ws.freeze_panes = "A2"


def autosize_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 28)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    csv_path = root / "analysis" / "Critical Comparison Report.csv"
    xlsx_path = root / "analysis" / "Critical Comparison Report.xlsx"

    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        source_headers = reader.fieldnames or []

    ordered_headers = reorder_columns(source_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    ws.append(ordered_headers)

    for row in rows:
        out = []
        for header in ordered_headers:
            raw = row.get(header, "")
            if header == "scenario":
                out.append(raw)
                continue
            try:
                out.append(float(raw))
            except (TypeError, ValueError):
                out.append(raw)
        ws.append(out)

    apply_styles(ws, ordered_headers)
    autosize_columns(ws)

    try:
        wb.save(xlsx_path)
        print(f"Created: {xlsx_path}")
    except PermissionError:
        # If Excel has the file open, write a timestamped copy instead of failing.
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = xlsx_path.with_name(f"Critical Comparison Report_{ts}.xlsx")
        wb.save(alt_path)
        print(f"Target XLSX is locked. Created alternative file: {alt_path}")


if __name__ == "__main__":
    main()
